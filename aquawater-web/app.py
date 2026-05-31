"""AquaWater 调洪演算程序 Web 版 — Flask 主应用"""

import io
import os
import sys
import secrets
from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from reservoir import run_flood_routing

app = Flask(__name__)

# ---- 安全配置 ----
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 上传限制 16MB
app.config['PREFERRED_URL_SCHEME'] = 'https'

# ---- 速率限制（防滥用）----
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["300 per day", "60 per hour"],
    storage_uri="memory://",
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 示例数据文件映射
SAMPLE_FILES = {
    'zv':  ('库容曲线.txt', '库容曲线_示例.txt'),
    'zq':  ('泄流.txt', '泄流曲线_示例.txt'),
    'flood': (os.path.join('sample', '洪水过程线.txt'), '洪水过程线_示例.txt'),
}


def parse_data_file(file_content):
    """解析数据文件，支持 Tab/逗号/空格 分隔，返回 [(col1, col2), ...]"""
    results = []
    for line in file_content.splitlines():
        line = line.strip()
        if not line:
            continue
        # 尝试用制表符、逗号、空格分隔
        parts = line.replace('\t', ' ').replace(',', ' ').split()
        if len(parts) >= 2:
            try:
                results.append((float(parts[0]), float(parts[1])))
            except ValueError:
                continue
    return results


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/preview', methods=['POST'])
def preview():
    """解析上传的文件，返回预览数据"""
    if 'file' not in request.files:
        return jsonify({'error': '未找到文件'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400
    try:
        content = file.read().decode('utf-8-sig')
        data = parse_data_file(content)
        if not data:
            return jsonify({'error': '未能解析到有效数据，请检查文件格式'}), 400
        return jsonify({
            'data': [{'col1': r[0], 'col2': r[1]} for r in data],
            'count': len(data),
        })
    except Exception as e:
        return jsonify({'error': f'文件解析失败: {str(e)}'}), 400


@app.route('/api/calculate', methods=['POST'])
@limiter.limit("30 per minute")
def calculate():
    """执行调洪演算"""
    try:
        body = request.get_json()
        zv_data = body.get('zv_data', [])
        zq_data = body.get('zq_data', [])
        flood_data = body.get('flood_data', [])
        z_start = float(body.get('z_start', 0))
        z0 = float(body.get('z0', 0))
        dt = float(body.get('dt', 0.1))
        method = int(body.get('method', 0))

        if not zv_data or not zq_data or not flood_data:
            return jsonify({'error': '请先导入完整的输入数据（库容曲线、泄流曲线、洪水过程线）'}), 400

        zv = [(r['col1'], r['col2']) for r in zv_data]
        zq = [(r['col1'], r['col2']) for r in zq_data]
        flood = [(r['col1'], r['col2']) for r in flood_data]

        result = run_flood_routing(zv, zq, flood, z_start, z0, dt, method)
        return jsonify(result)

    except ValueError as e:
        return jsonify({'error': f'参数格式错误: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'计算失败: {str(e)}'}), 500


@app.route('/api/export', methods=['POST'])
def export():
    """导出结果为 Excel 文件"""
    try:
        body = request.get_json()
        results = body.get('results', [])
        summary = body.get('summary', {})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '调洪成果'

        # 样式定义
        header_font = Font(name='宋体', bold=True, size=12)
        title_font = Font(name='宋体', bold=True, size=16)
        data_font = Font(name='宋体', size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # 标题行
        ws.merge_cells('A1:H1')
        ws['A1'] = '调洪演算成果表'
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        # 特征值
        ws.merge_cells('A3:H3')
        ws['A3'] = '特征值'
        ws['A3'].font = header_font

        summary_items = [
            ('起调水位(m)', summary.get('z_start', '')),
            ('起调库容(万m³)', summary.get('v_start', '')),
            ('调洪最高水位(m)', summary.get('z_max', '')),
            ('调洪最大库容(万m³)', summary.get('v_max', '')),
            ('调洪最大下泄流量(m³/s)', summary.get('q_max', '')),
            ('滞洪库容(万m³)', summary.get('v_retention', '')),
        ]
        for i, (label, value) in enumerate(summary_items):
            ws.cell(row=4 + i, column=1, value=label).font = data_font
            ws.cell(row=4 + i, column=2, value=value).font = data_font

        # 结果表头
        start_row = 11
        headers = ['序号', '时间(h)', '时段平均入库(m³/s)', '时段平均出库(m³/s)',
                   '时段初水位(m)', '时段末水位(m)', '时段初库容(万m³)', '时段末库容(万m³)']
        for j, h in enumerate(headers):
            cell = ws.cell(row=start_row, column=j + 1, value=h)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # 结果数据
        for i, row in enumerate(results):
            values = [
                i + 1,
                row.get('time', ''),
                row.get('inflow_avg', ''),
                row.get('outflow_avg', ''),
                row.get('z_begin', ''),
                row.get('z_end', ''),
                row.get('v_begin', ''),
                row.get('v_end', ''),
            ]
            for j, val in enumerate(values):
                cell = ws.cell(row=start_row + 1 + i, column=j + 1, value=val)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border

        # 调整列宽
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='调洪成果.xlsx',
        )

    except Exception as e:
        return jsonify({'error': f'导出失败: {str(e)}'}), 500


@app.route('/api/download-sample/<file_type>')
def download_sample(file_type):
    """下载示例数据文件"""
    if file_type not in SAMPLE_FILES:
        return jsonify({'error': '未知的示例文件类型'}), 400

    file_path, download_name = SAMPLE_FILES[file_type]
    full_path = os.path.join(BASE_DIR, file_path)

    if not os.path.exists(full_path):
        return jsonify({'error': '示例文件不存在'}), 404

    return send_file(
        full_path,
        mimetype='text/plain; charset=utf-8',
        as_attachment=True,
        download_name=download_name,
    )


if __name__ == '__main__':
    from waitress import serve
    port = int(os.environ.get('PORT', 8090))
    print(f'  AquaWater 调洪演算服务已启动')
    print(f'  访问地址: http://localhost:{port}')
    print(f'  按 Ctrl+C 停止服务')
    serve(app, host='0.0.0.0', port=port)
